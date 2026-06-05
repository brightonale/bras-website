import openpyxl
import json
import re
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb_path = r"I:\My Drive\BRAS Backend Data\BRAS_Master_Rankings.xlsx"
output_dir = r"C:\Users\harry\Documents\antigravity\fearless-nobel\src\data"

if not os.path.exists(output_dir):
    os.makedirs(output_dir)

wb = openpyxl.load_workbook(wb_path, data_only=True)

def get_academic_year(date_str):
    if not date_str:
        return "All"
    date_str = str(date_str).strip()
    if "/" in date_str:
        parts = date_str.split("/")
        if len(parts) == 2:
            return f"{parts[0][-2:]}/{parts[1][-2:]}"
    
    # Match strings like "23 Oct 2025" or "18 Dec 2025"
    year_match = re.search(r'\b(202\d)\b', date_str)
    if year_match:
        year = int(year_match.group(1))
        # Default to academic year: if month is Jan-Aug, it's year-1/year. If Sep-Dec, it's year/year+1
        month = 10 # default to Oct
        for m_name in ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug"]:
            if m_name in date_str.lower():
                month = 1
                break
        if month >= 9:
            return f"{str(year)[-2:]}/{str(year+1)[-2:]}"
        else:
            return f"{str(year-1)[-2:]}/{str(year)[-2:]}"
    return "25/26" # default fallback for current season

# 1. Pub Leaderboard
pubs = []
if "Pub Leaderboard" in wb.sheetnames:
    ws = wb["Pub Leaderboard"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            score_val = 0.0
            try:
                score_val = float(row[5]) if row[5] is not None else 0.0
            except:
                pass
            
            pubs.append({
                "rank": row[0],
                "pub": row[1],
                "pint": row[2] or "",
                "brewery": row[3] or "",
                "date": str(row[4]) if row[4] is not None else "",
                "academicYear": get_academic_year(row[4]),
                "score": score_val,
                "ratingsCount": str(row[6]) if row[6] is not None else "0"
            })

# 2. Pubs Checklist
checklist = []
if "Pubs Checklist" in wb.sheetnames:
    ws = wb["Pubs Checklist"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            checklist.append({
                "status": row[0] or "Not Visited",
                "name": row[1],
                "comment": row[2] or ""
            })

# 3. Member × Pub Matrix
matrix_data = {"pubs": [], "rows": []}
if "Member × Pub Matrix" in wb.sheetnames:
    ws = wb["Member × Pub Matrix"]
    headers = [col for col in next(ws.iter_rows(max_row=1, values_only=True))]
    pub_headers = headers[1:]
    matrix_data["pubs"] = pub_headers
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            member_ratings = {}
            for col_idx, pub_name in enumerate(pub_headers, 1):
                val = row[col_idx]
                if val == '—' or val is None:
                    member_ratings[pub_name] = None
                else:
                    try:
                        member_ratings[pub_name] = float(val)
                    except:
                        member_ratings[pub_name] = None
            
            matrix_data["rows"].append({
                "member": row[0],
                "ratings": member_ratings
            })

# Map matrix ratings to members lists for easier display in profiles
member_matrix_map = {r["member"]: r["ratings"] for r in matrix_data["rows"]}

# 4. Member Leaderboard
members = []
if "Member Leaderboard" in wb.sheetnames:
    ws = wb["Member Leaderboard"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            m_name = row[1]
            ratings_dict = member_matrix_map.get(m_name, {})
            visited_list = []
            for p_name, score in ratings_dict.items():
                if score is not None:
                    visited_list.append({"pub": p_name, "score": score})
            
            members.append({
                "rank": row[0],
                "name": m_name,
                "pubsVisited": int(row[2]) if row[2] is not None else 0,
                "totalRatings": int(row[3]) if row[3] is not None else 0,
                "avgScoreGiven": float(row[4]) if row[4] is not None else 0.0,
                "highestGiven": float(row[5]) if row[5] is not None else 0.0,
                "lowestGiven": float(row[6]) if row[6] is not None else 0.0,
                "visitedPubs": visited_list
            })

# 5. Visit Timeline
timeline = []
if "Visit Timeline" in wb.sheetnames:
    ws = wb["Visit Timeline"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            score_val = 0.0
            try:
                score_val = float(row[3]) if row[3] is not None else 0.0
            except:
                pass
            
            timeline.append({
                "number": row[0],
                "date": str(row[1]) if row[1] is not None else "",
                "academicYear": get_academic_year(row[1]),
                "pub": row[2],
                "avgScore": score_val,
                "attendees": int(row[4]) if row[4] is not None else 0,
                "membersPresent": str(row[5]) if row[5] is not None else ""
            })

# Write JSON Files
with open(os.path.join(output_dir, "pubs.json"), "w", encoding="utf-8") as f:
    json.dump(pubs, f, ensure_ascii=False, indent=2)

with open(os.path.join(output_dir, "checklist.json"), "w", encoding="utf-8") as f:
    json.dump(checklist, f, ensure_ascii=False, indent=2)

with open(os.path.join(output_dir, "matrix.json"), "w", encoding="utf-8") as f:
    json.dump(matrix_data, f, ensure_ascii=False, indent=2)

with open(os.path.join(output_dir, "members.json"), "w", encoding="utf-8") as f:
    json.dump(members, f, ensure_ascii=False, indent=2)

with open(os.path.join(output_dir, "timeline.json"), "w", encoding="utf-8") as f:
    json.dump(timeline, f, ensure_ascii=False, indent=2)

print("Data compiled successfully!")
print(f"Pubs: {len(pubs)}, Checklist items: {len(checklist)}, Members: {len(members)}, Timeline events: {len(timeline)}")
